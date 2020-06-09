import sys
import os

sys.path.append(os.path.abspath('./').replace('\\', '/'))

from selenium.webdriver.common.by import By
import openpyxl
from base import Page
from time import sleep
from loginPage import login
from selenium import webdriver


class posting(Page):
    """用户发帖界面类"""

    def __init__(self, s_driver, row):
        super().__init__(s_driver)
        self.url = '/forum.php?mod=post&action=newthread&fid=138'
        if 'test_case' in os.path.abspath('./'):
            path = os.path.abspath('./').split('test_case')[0]
            path = str(path) + '\\data\\posts.xlsx'
        else:
            path = os.path.abspath('./')+"\\bbs\\data\\posts.xlsx"
        wb = openpyxl.load_workbook(path)
        self.ws = wb.active
        self.subject_input_loc = (By.ID, "subject")
        self.chinese_name_input_loc = (By.ID, "typeoption_chinapluginname")
        self.map_class2_box_loc = (By.CSS_SELECTOR, "#select_map_type > ul:nth-child(1) > li:nth-child(2) > label:nth-child(1) > input:nth-child(1)")
        self.map_class6_box_loc = (By.CSS_SELECTOR, "#select_map_type > ul:nth-child(1) > li:nth-child(5) > label:nth-child(1) > input:nth-child(1)")
        self.map_log_loc = (By.CSS_SELECTOR, '#select_mapfile_type li:nth-child(1) > label')
        self.carry_class_loc = (By.CSS_SELECTOR, '#select_dfrom > ul:nth-child(1) > li:nth-child(3) > label:nth-child(1) > input:nth-child(1)')
        self.carry_source_url_loc = (By.CSS_SELECTOR, '#typeoption_original_url')
        self.download_url_loc = (By.CSS_SELECTOR, '#typeoption_download_src')
        self.text_intro_frame_loc = (By.CSS_SELECTOR, '#e_iframe')
        self.text_intro_loc = (By.TAG_NAME, 'body')
        self.submit_button_loc = (By.CSS_SELECTOR, 'button.pn:nth-child(2)')

        self.i = row  # 第三条记录

        self.chinese_name = self.ws['C' + str(self.i)].value
        self.map_class = self.ws['F' + str(self.i)].value
        self.map_log = self.ws['D' + str(self.i)].value
        self.carry_class = self.ws['E' + str(self.i)].value
        self.carry_source_url = self.ws['A' + str(self.i)].value
        self.download_url = self.ws['B' + str(self.i)].value
        self.text_intro = self.ws['G' + str(self.i)].value

    def set_subject(self):
        self.find_element(*self.subject_input_loc).send_keys(str(self.chinese_name)+'地图')

    def set_chinese_name(self):
        self.find_element(*self.chinese_name_input_loc).send_keys(str(self.chinese_name))

    def set_map_class(self):
        if self.map_class == '解密':
            self.find_element(*self.map_class2_box_loc).click()
        elif self.map_class == '风景':
            self.find_element(*self.map_class6_box_loc).click()

    def set_map_log(self):
        self.find_element(*self.map_log_loc).click()

    def set_carry_class(self):
        self.find_element(*self.carry_class_loc).click()

    def set_source_url(self):
        self.find_element(*self.carry_source_url_loc).send_keys(str(self.carry_source_url))

    def set_download_url(self):
        self.find_element(*self.download_url_loc).send_keys(str(self.download_url))

    def set_text_intro(self):
        self.script("arguments[0].scrollIntoView(false);", self.find_element(*self.submit_button_loc))
        self.driver.switch_to.frame(self.find_element(*self.text_intro_frame_loc))
        self.find_element(*self.text_intro_loc).send_keys(str(self.text_intro))

    def submit_post(self):
        self.driver.switch_to.default_content()
        # self.find_element(*self.submit_button_loc).click()
        self.find_element(*self.submit_button_loc).click()

    def user_posting(self):
        """统一发帖入口"""
        self.open()
        self.set_subject()
        self.set_chinese_name()
        self.set_map_class()
        self.set_carry_class()
        self.set_source_url()
        self.set_download_url()
        self.set_text_intro()
        self.submit_post()


if __name__ == '__main__':
    username = 'minecraft_ak'
    password = '2438675akkl'
    driver = webdriver.Firefox()
    lo = login(driver)
    pt = posting(driver, 3)
    print(pt.carry_class)
    lo.user_login(username, password)
    pt.user_posting()
    sleep(3)
    lo.logout()
